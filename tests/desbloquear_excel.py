import os
import sys
import zipfile
import tempfile
import shutil
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import webbrowser
from datetime import datetime
import openpyxl
from openpyxl.worksheet.protection import SheetProtection

def desbloquear_excel(arquivo_excel):
    """
    Remove toda proteção de um arquivo Excel (workbook e planilhas)
    manipulando diretamente os arquivos XML internos.
    """
    if not os.path.exists(arquivo_excel):
        print(f"Erro: O arquivo {arquivo_excel} não existe.")
        return False
    
    print(f"Desbloqueando arquivo: {arquivo_excel}")
    
    # Cria um diretório temporário
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Copia o arquivo para o diretório temporário
        temp_file = os.path.join(temp_dir, "temp.xlsx")
        shutil.copy2(arquivo_excel, temp_file)
        
        # Renomeia para .zip (Excel é um arquivo ZIP)
        zip_file = os.path.join(temp_dir, "temp.zip")
        os.rename(temp_file, zip_file)
        
        # Extrai o conteúdo
        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Modifica os arquivos XML para remover proteção
        
        # 1. Remover proteção do workbook
        workbook_file = os.path.join(temp_dir, 'xl', 'workbook.xml')
        if os.path.exists(workbook_file):
            print("Removendo proteção do workbook...")
            try:
                with open(workbook_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                if '<workbookProtection' in content:
                    start_idx = content.find('<workbookProtection')
                    if start_idx >= 0:
                        end_idx = content.find('/>', start_idx)
                        if end_idx >= 0:
                            end_idx += 2
                            content = content[:start_idx] + content[end_idx:]
                            print("Proteção do workbook removida")
                
                with open(workbook_file, 'w', encoding='utf-8') as f:
                    f.write(content)
                
            except Exception as e:
                print(f"Aviso ao modificar workbook.xml: {str(e)}")
        
        # 2. Remover proteção das planilhas
        worksheets_dir = os.path.join(temp_dir, 'xl', 'worksheets')
        if os.path.exists(worksheets_dir):
            print("Removendo proteção das planilhas...")
            for sheet_file in os.listdir(worksheets_dir):
                if sheet_file.startswith('sheet') and sheet_file.endswith('.xml'):
                    sheet_path = os.path.join(worksheets_dir, sheet_file)
                    try:
                        with open(sheet_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                        
                        if '<sheetProtection' in content:
                            start_idx = content.find('<sheetProtection')
                            if start_idx >= 0:
                                end_idx = content.find('/>', start_idx)
                                if end_idx >= 0:
                                    end_idx += 2
                                    content = content[:start_idx] + content[end_idx:]
                                    print(f"Proteção removida da planilha: {sheet_file}")
                        
                        with open(sheet_path, 'w', encoding='utf-8') as f:
                            f.write(content)
                        
                    except Exception as e:
                        print(f"Aviso ao modificar {sheet_file}: {str(e)}")
        
        # Recria o arquivo ZIP
        print("Recriando arquivo Excel...")
        new_zip_file = os.path.join(temp_dir, "new.xlsx")
        with zipfile.ZipFile(new_zip_file, 'w') as zipf:
            for root_dir, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file not in ["new.xlsx", "temp.xlsx", "temp.zip"]:
                        file_path = os.path.join(root_dir, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        zipf.write(file_path, arcname)
        
        # Substitui o arquivo original
        shutil.copy2(new_zip_file, arquivo_excel)
        print(f"Arquivo desbloqueado com sucesso!")
        
        return True
    
    except Exception as e:
        print(f"Erro ao desbloquear arquivo: {str(e)}")
        return False
    
    finally:
        # Limpa os arquivos temporários
        try:
            shutil.rmtree(temp_dir)
        except:
            pass

def bloquear_excel(arquivo_excel, senha, range_celulas="A1:AI107"):
    """
    Bloqueia a edição de células específicas em todas as planilhas de um arquivo Excel.
    
    Parâmetros:
    - arquivo_excel: caminho do arquivo Excel
    - senha: senha para proteção
    - range_celulas: range de células a serem bloqueadas (padrão: A1:AI107)
    """
    if not os.path.exists(arquivo_excel):
        print(f"Erro: O arquivo {arquivo_excel} não existe.")
        return False
    
    try:
        # Carrega o workbook
        wb = openpyxl.load_workbook(arquivo_excel)
        
        # Processa cada planilha
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Processando planilha: {sheet_name}")
            
            # Primeiro, desbloqueia todas as células
            for row in sheet.iter_rows():
                for cell in row:
                    cell.protection = openpyxl.styles.Protection(locked=False)
            
            # Bloqueia o range específico
            start_cell, end_cell = range_celulas.split(':')
            start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
            end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)
            
            start_col_idx = openpyxl.utils.column_index_from_string(start_col)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col)
            
            # Bloqueia apenas as células dentro do range especificado
            for row in range(start_row, end_row + 1):
                for col in range(start_col_idx, end_col_idx + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.protection = openpyxl.styles.Protection(locked=True)
            
            # Protege a planilha com senha
            sheet.protection = SheetProtection(
                sheet=True,
                password=senha,
                selectLockedCells=True,
                selectUnlockedCells=True,
                formatCells=False,
                formatColumns=False,
                formatRows=False,
                insertColumns=False,
                insertRows=False,
                insertHyperlinks=False,
                deleteColumns=False,
                deleteRows=False,
                sort=False,
                autoFilter=False,
                pivotTables=False
            )
            
            print(f"Planilha {sheet_name} protegida com sucesso")
        
        # Salva diretamente no arquivo original
        wb.save(arquivo_excel)
        print(f"Arquivo bloqueado com sucesso!")
        return True
        
    except Exception as e:
        print(f"Erro ao bloquear arquivo: {str(e)}")
        return False

class DesbloquearExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gerenciador de Proteção Excel - v2.0")
        
        # Configura o tamanho inicial e mínimo
        self.root.geometry("800x600")
        self.root.minsize(700, 500)
        
        # Configura o ícone (se disponível)
        try:
            self.root.iconbitmap("excel_icon.ico")
        except:
            pass
        
        # Configura o tema
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Cores e estilos personalizados
        self.style.configure("Header.TLabel", 
                           font=('Segoe UI', 16, 'bold'), 
                           foreground="#1e3d59")
        self.style.configure("Status.TLabel", 
                           font=('Segoe UI', 9),
                           background="#f5f5f5")
        self.style.configure("Action.TButton",
                           font=('Segoe UI', 10, 'bold'),
                           padding=10)
        
        # Frame principal com padding
        self.main_frame = ttk.Frame(root, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Cabeçalho
        header_frame = ttk.Frame(self.main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(header_frame, 
                 text="Gerenciador de Proteção Excel em Lote", 
                 style="Header.TLabel").pack(side=tk.LEFT)
        
        # Botão de ajuda
        ttk.Button(header_frame, 
                  text="?", 
                  width=3,
                  command=self.mostrar_ajuda).pack(side=tk.RIGHT)
        
        # Frame de modo
        mode_frame = ttk.LabelFrame(self.main_frame, text="Modo de Operação", padding=10)
        mode_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Radio buttons para selecionar o modo
        self.modo_var = tk.StringVar(value="desbloquear")
        ttk.Radiobutton(mode_frame, text="Desbloquear Excel", 
                       variable=self.modo_var, value="desbloquear",
                       command=self.atualizar_interface).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(mode_frame, text="Bloquear Excel", 
                       variable=self.modo_var, value="bloquear",
                       command=self.atualizar_interface).pack(side=tk.LEFT, padx=10)
        
        # Frame para seleção de diretório
        self.dir_frame = ttk.LabelFrame(self.main_frame, text="Seleção de Diretório", padding=10)
        self.dir_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Grid para organizar os elementos
        self.dir_frame.columnconfigure(1, weight=1)
        
        # Diretório de entrada
        ttk.Label(self.dir_frame, text="Diretório:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.entrada_diretorio = ttk.Entry(self.dir_frame)
        self.entrada_diretorio.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(self.dir_frame, text="Procurar...", 
                  command=self.selecionar_diretorio).grid(row=0, column=2, padx=5, pady=5)
        
        # Frame para opções de bloqueio
        self.block_frame = ttk.LabelFrame(self.main_frame, text="Opções de Bloqueio", padding=10)
        
        # Grid para organizar os elementos de bloqueio
        self.block_frame.columnconfigure(1, weight=1)
        
        # Campo para senha
        ttk.Label(self.block_frame, text="Senha:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.entrada_senha = ttk.Entry(self.block_frame, show="*")
        self.entrada_senha.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        
        # Campo para range de células
        ttk.Label(self.block_frame, text="Range:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.entrada_range = ttk.Entry(self.block_frame)
        self.entrada_range.insert(0, "A1:AI107")
        self.entrada_range.grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)
        
        # Frame de opções gerais
        self.options_frame = ttk.LabelFrame(self.main_frame, text="Opções Gerais", padding=10)
        self.options_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Checkbox para incluir subdiretórios
        self.subdirs_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(self.options_frame, 
                       text="Incluir subdiretórios", 
                       variable=self.subdirs_var).pack(anchor=tk.W)
        
        # Frame de ações
        actions_frame = ttk.Frame(self.main_frame)
        actions_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Botão de ação principal
        self.btn_principal = ttk.Button(actions_frame, 
                                      text="Processar Arquivos", 
                                      style="Action.TButton",
                                      command=self.executar_acao)
        self.btn_principal.pack(pady=5)
        
        # Área de log com título e contador
        log_frame = ttk.Frame(self.main_frame)
        log_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(log_frame, text="Log de Operações:", anchor=tk.W).pack(side=tk.LEFT)
        self.contador_label = ttk.Label(log_frame, text="Arquivos processados: 0/0")
        self.contador_label.pack(side=tk.RIGHT)
        
        # Frame para o log com borda
        self.log_frame = ttk.Frame(self.main_frame, relief=tk.SUNKEN, borderwidth=1)
        self.log_frame.pack(fill=tk.BOTH, expand=True)
        
        # Configuração do log com scrollbar
        self.log_text = tk.Text(self.log_frame, 
                               wrap=tk.WORD, 
                               font=('Consolas', 9),
                               background="#f8f8f8")
        self.scrollbar = ttk.Scrollbar(self.log_frame, 
                                     orient=tk.VERTICAL, 
                                     command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=self.scrollbar.set)
        
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Barra de status
        self.status_var = tk.StringVar(value="Pronto")
        self.status_bar = ttk.Label(root, 
                                  textvariable=self.status_var,
                                  style="Status.TLabel",
                                  relief=tk.SUNKEN,
                                  anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Configura o menu
        self.criar_menu()
        
        # Inicia com uma mensagem de boas-vindas
        self.adicionar_log("Bem-vindo ao Gerenciador de Proteção Excel em Lote!")
        self.adicionar_log("Selecione um diretório para começar.")
        
        # Atualiza a interface inicial
        self.atualizar_interface()
        
        # Variáveis para controle de processamento
        self.arquivos_total = 0
        self.arquivos_processados = 0
        self.processamento_ativo = False
    
    def criar_menu(self):
        """Cria a barra de menu"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # Menu Arquivo
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Arquivo", menu=file_menu)
        file_menu.add_command(label="Selecionar Diretório...", 
                            command=self.selecionar_diretorio)
        file_menu.add_separator()
        file_menu.add_command(label="Sair", command=self.root.quit)
        
        # Menu Opções
        options_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Opções", menu=options_menu)
        options_menu.add_checkbutton(label="Incluir Subdiretórios", 
                                   variable=self.subdirs_var)
        
        # Menu Ajuda
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ajuda", menu=help_menu)
        help_menu.add_command(label="Documentação", command=self.mostrar_ajuda)
        help_menu.add_command(label="Sobre", command=self.mostrar_sobre)
    
    def mostrar_ajuda(self):
        """Mostra a janela de ajuda"""
        ajuda = tk.Toplevel(self.root)
        ajuda.title("Ajuda - Desbloqueador de Excel")
        ajuda.geometry("500x400")
        ajuda.transient(self.root)
        ajuda.grab_set()
        
        # Texto de ajuda
        texto = """
Como usar o Desbloqueador de Excel:

1. Selecione o arquivo Excel protegido:
   - Clique em "Procurar..." ou use o menu Arquivo > Abrir Excel
   - Escolha o arquivo .xlsx que deseja desbloquear

2. Escolha onde salvar o arquivo desbloqueado:
   - Por padrão, será salvo com "_desbloqueado" no nome
   - Você pode escolher outro local/nome clicando em "Procurar..."

3. Opções:
   - Backup automático: cria uma cópia do arquivo original
     antes de fazer alterações (recomendado)

4. Clique em "Desbloquear Excel":
   - O processo será iniciado e você poderá acompanhar
     o progresso na área de log
   - Aguarde até a conclusão do processo

5. Verificação:
   - Após o processo, verifique se o arquivo foi desbloqueado
   - O novo arquivo estará salvo no local especificado

Observações:
- O processo é seguro e não altera o arquivo original
- Recomenda-se manter o backup ativado
- Em caso de erro, verifique o log para mais detalhes
"""
        
        text_widget = tk.Text(ajuda, wrap=tk.WORD, padx=10, pady=10)
        text_widget.pack(fill=tk.BOTH, expand=True)
        text_widget.insert(tk.END, texto)
        text_widget.config(state=tk.DISABLED)
        
        ttk.Button(ajuda, text="Fechar", command=ajuda.destroy).pack(pady=10)
    
    def mostrar_sobre(self):
        """Mostra a janela Sobre"""
        messagebox.showinfo(
            "Sobre - Desbloqueador de Excel",
            "Desbloqueador de Excel v2.0\n\n"
            "Desenvolvido para facilitar o desbloqueio\n"
            "de arquivos Excel protegidos.\n\n"
            "© 2024 - Todos os direitos reservados"
        )
    
    def selecionar_diretorio(self):
        """Seleciona o diretório de entrada"""
        diretorio = filedialog.askdirectory(
            title="Selecione o diretório com arquivos Excel"
        )
        if diretorio:
            self.entrada_diretorio.delete(0, tk.END)
            self.entrada_diretorio.insert(0, diretorio)
            self.adicionar_log(f"Diretório selecionado: {diretorio}")
            self.contar_arquivos_excel()
    
    def contar_arquivos_excel(self):
        """Conta quantos arquivos Excel existem no diretório"""
        diretorio = self.entrada_diretorio.get()
        if not diretorio:
            return
        
        total = 0
        if self.subdirs_var.get():
            for root, _, files in os.walk(diretorio):
                total += sum(1 for f in files if f.endswith('.xlsx'))
        else:
            total = sum(1 for f in os.listdir(diretorio) if f.endswith('.xlsx'))
        
        self.arquivos_total = total
        self.arquivos_processados = 0
        self.atualizar_contador()
        self.adicionar_log(f"Encontrados {total} arquivos Excel no diretório")
    
    def atualizar_contador(self):
        """Atualiza o contador de arquivos processados"""
        self.contador_label.config(
            text=f"Arquivos processados: {self.arquivos_processados}/{self.arquivos_total}"
        )
    
    def processar_arquivo(self, arquivo):
        """Processa um único arquivo Excel"""
        modo = self.modo_var.get()
        
        try:
            if modo == "bloquear":
                senha = self.entrada_senha.get()
                range_celulas = self.entrada_range.get()
                sucesso = bloquear_excel(arquivo, senha, range_celulas)
            else:
                sucesso = desbloquear_excel(arquivo)
            
            if sucesso:
                self.adicionar_log(f"✓ Arquivo processado com sucesso: {os.path.basename(arquivo)}")
            else:
                self.adicionar_log(f"✗ Falha ao processar: {os.path.basename(arquivo)}")
            
            return sucesso
            
        except Exception as e:
            self.adicionar_log(f"✗ Erro ao processar {os.path.basename(arquivo)}: {str(e)}")
            return False
    
    def processar_diretorio(self):
        """Processa todos os arquivos Excel no diretório"""
        diretorio = self.entrada_diretorio.get()
        modo = self.modo_var.get()
        
        if not diretorio or not os.path.isdir(diretorio):
            messagebox.showerror("Erro", "Por favor, selecione um diretório válido.")
            return
        
        if modo == "bloquear" and not self.entrada_senha.get():
            messagebox.showerror("Erro", "Por favor, defina uma senha para o bloqueio.")
            return
        
        self.processamento_ativo = True
        self.btn_principal.config(state=tk.DISABLED)
        self.log_text.delete(1.0, tk.END)
        self.adicionar_log(f"Iniciando processamento em lote no diretório: {diretorio}")
        
        try:
            arquivos_excel = []
            if self.subdirs_var.get():
                for root, _, files in os.walk(diretorio):
                    for file in files:
                        if file.endswith('.xlsx'):
                            arquivos_excel.append(os.path.join(root, file))
            else:
                arquivos_excel = [os.path.join(diretorio, f) for f in os.listdir(diretorio) 
                                if f.endswith('.xlsx')]
            
            self.arquivos_total = len(arquivos_excel)
            self.arquivos_processados = 0
            self.atualizar_contador()
            
            for arquivo in arquivos_excel:
                if not self.processamento_ativo:
                    break
                
                self.processar_arquivo(arquivo)
                self.arquivos_processados += 1
                self.atualizar_contador()
                self.root.update_idletasks()
            
            if self.processamento_ativo:
                self.adicionar_log("\nProcessamento em lote concluído!")
                messagebox.showinfo("Sucesso", 
                                  f"Processamento concluído!\n"
                                  f"Total de arquivos processados: {self.arquivos_processados}")
        
        except Exception as e:
            self.adicionar_log(f"Erro durante o processamento em lote: {str(e)}")
            messagebox.showerror("Erro", f"Ocorreu um erro durante o processamento:\n{str(e)}")
        
        finally:
            self.processamento_ativo = False
            self.btn_principal.config(state=tk.NORMAL)
            self.status_var.set("Pronto")
    
    def executar_acao(self):
        """Inicia o processamento em lote em uma thread separada"""
        threading.Thread(target=self.processar_diretorio, daemon=True).start()
    
    def adicionar_log(self, mensagem):
        """Adiciona mensagem ao log com timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {mensagem}\n")
        self.log_text.see(tk.END)
        self.status_var.set(mensagem)
        self.root.update_idletasks()
    
    def atualizar_interface(self):
        """Atualiza a interface baseado no modo selecionado"""
        modo = self.modo_var.get()
        
        if modo == "bloquear":
            self.block_frame.pack(after=self.dir_frame, fill=tk.X, pady=(0, 10))
            self.btn_principal.configure(text="Bloquear Arquivos")
            self.root.title("Gerenciador de Proteção Excel - Modo: Bloqueio")
            self.adicionar_log("Modo alterado para: Bloqueio")
        else:
            self.block_frame.pack_forget()
            self.btn_principal.configure(text="Desbloquear Arquivos")
            self.root.title("Gerenciador de Proteção Excel - Modo: Desbloqueio")
            self.adicionar_log("Modo alterado para: Desbloqueio")
        
        # Atualiza a contagem de arquivos se houver um diretório selecionado
        if self.entrada_diretorio.get():
            self.contar_arquivos_excel()

def main():
    # Verifica se deve usar a interface gráfica ou linha de comando
    if len(sys.argv) > 1 and sys.argv[1] == "--nogui":
        # Modo linha de comando
        if len(sys.argv) > 2:
            arquivo_excel = sys.argv[2]
            arquivo_saida = sys.argv[3] if len(sys.argv) > 3 else None
        else:
            arquivo_excel = input("Digite o caminho completo para o arquivo Excel protegido: ")
            arquivo_saida = input("Digite o caminho para salvar o arquivo desbloqueado (ou deixe em branco para usar o padrão): ")
            if not arquivo_saida:
                arquivo_saida = None
        
        desbloquear_excel(arquivo_excel)
    else:
        # Modo interface gráfica
        root = tk.Tk()
        app = DesbloquearExcelApp(root)
        root.mainloop()

if __name__ == "__main__":
    main() 